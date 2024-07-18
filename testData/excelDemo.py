import openpyxl
## location
book = openpyxl.load_workbook("C:\\Users\\RameshKaramthot\\PycharmProjects\\workshop-pytest\\testData\\testdatafilexcel.xlsx")
## focus on sheet which is active
sheet = book.active
## and cell
cell = sheet.cell(row=2,column=2)
print(cell.value)
sheet.cell(row=5,column=1).value = "testcase4"
print(sheet.cell(row=5,column=1).value)

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A4'].value)

for i in range(1, sheet.max_row+1):    # to get rows

        for j in range(2, sheet.max_column+1):  # to get columns
            print(sheet.cell(row=i, column=j).value)


