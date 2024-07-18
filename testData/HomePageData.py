import openpyxl


class HomePageData:

    test_homepage_data = [{"firstname":"Python", "email":"pytest2233@gmail.com", "password":"password", "gender":"Male"}]


    @staticmethod
    def getTestData(test_case_name):

        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\RameshKaramthot\\PycharmProjects\\workshop-pytest\\testData\\testdatafilexcel.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    #print(sheet.cell(row=i, column=j).value)
                    # dict["name"] = "testcase3"     to add the values in empty dictionary
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]